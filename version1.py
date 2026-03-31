############################    STORAGE Algorithm     ###########################
# Version 1

import pandas as pd
from openpyxl import Workbook
import os

# --- CONFIGURATION ---
base_path = os.path.dirname(__file__)
input_path = os.path.join(base_path, 'walmartSales.xlsx')
output_path = os.path.join(base_path, 'walmart_version1.xlsx')

wb = Workbook()
ws = wb.active

try:
    df = pd.read_excel(input_path, sheet_name='Walmart sales')
except Exception as e:
    print(f"Error loading file: {e}. Please ensure the file is in the same directory.")
    exit()

# Transform data into a dictionary for easier processing
d = {str(i[0]): str(i[1]).split() for i in df.values}

for g in d:
    for k, f in enumerate(d[g]):
        d[g][k] = int(f)

def letters(x):
    y = str(x)
    mapping = 'zabcdefghi'
    if len(y) == 1: 
        return mapping[int(y)]
    else:
        return ''.join(mapping[int(i)] for i in y)

def InvLetters(x):
    mapping = 'zabcdefghi'
    n = ''
    for i in x:
        n += str(mapping.find(i))
    return int(n)

q = 1
for b in d:
    x = d[b]
    last = ''
    kk = 0
    
    while kk < len(x):
        k = kk + 1
        # Cicle for counting consecutive identical values
        while k < len(x) and x[k] == x[kk]:
            k += 1
        
        count = k - kk
        valore = x[kk]
        
        # Logic for creating the compressed segment: if count > 1 or value > 9, we add the letter code, otherwise we just add the value
        segmento = ""
        if count > 1 or valore > 9:
            segmento = f"{valore}{letters(count)}"
        else:
            segmento = f"{valore}"
            
        if last == "":
            last = segmento
        else:
            # We check if we can concatenate without space (if last ends with a digit and segmento starts with a digit)
            if last[-1].isdigit() and segmento[0].isdigit():
                last += f" {segmento}"
            else:
                last += segmento
        
        kk = k # We move to the next segment

    ws[f'A{q}'] = b
    ws[f'B{q}'] = last
    q += 1

    c = []
    g_idx = 0
    temp_last = last
    
    # Parse the compressed string to reconstruct the original list of numbers
    import re
    # Find all segments of the form "number + optional letters"
    parts = re.findall(r'\d+[a-z]*', temp_last)
    
    for p in parts:
        # Separate the numeric part from the letter part
        num_part = "".join(filter(str.isdigit, p))
        let_part = "".join(filter(str.isalpha, p))
        
        val = int(num_part)
        freq = InvLetters(let_part) if let_part else 1
        c.extend([val] * freq)

    if x != c:
        print(f"Error in reconstruction for ID {b}. Original: {x}, Reconstructed: {c}")

wb.save(output_path)
print(f"Process completed. File saved in: {output_path}")