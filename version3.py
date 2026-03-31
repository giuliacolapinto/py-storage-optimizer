############################    STORAGE Algorithm     ###########################
# Version 3

import pandas as pd
from openpyxl import Workbook
from collections import Counter
import os

# --- CONFIGURATION ---
input_path = '/Users/giuliacolapinto/Library/Mobile Documents/com~apple~CloudDocs/Programmazione/optimizeSpace/walmartSales.xlsx'
output_path = "/Users/giuliacolapinto/Library/Mobile Documents/com~apple~CloudDocs/Programmazione/optimizeSpace/walmart_version3.xlsx"

# --- DATA LOADING ---
try:
    df = pd.read_excel(input_path)
except Exception as e:
    print(f"Error loading file: {e}")
    exit()

all_rows = [str(row[1]).split() for row in df.values]

# --- GLOBAL TOKENIZATION DICTIONARY ---
# Identify sequences of 3 numbers that repeat most frequently
patterns = []
for row in all_rows:
    for i in range(len(row) - 2):
        patterns.append(",".join(row[i:i+3]))

# Map the top 52 patterns to letters (A-Z, a-z)
top_patterns = [p[0] for p in Counter(patterns).most_common(52)]
token_dict = {pattern: chr(65 + i) if i < 26 else chr(71 + i) for i, pattern in enumerate(top_patterns)}

# --- COMPRESSION ---
wb = Workbook()
ws = wb.active
ws.title = "T"

for index, row in enumerate(df.values):
    id_v = row[0]
    row_str = ",".join(str(row[1]).split())
    
    # Replace frequent patterns with dictionary tokens
    for pattern, token in token_dict.items():
        row_str = row_str.replace(pattern, token)
    
    # Remove remaining commas to minimize space
    compressed = row_str.replace(",", "")
    
    ws.cell(row=index + 1, column=1, value=id_v)
    ws.cell(row=index + 1, column=2, value=compressed)

# --- SAVE ---
wb.save(output_path)

file_size = os.path.getsize(output_path) / 1024
print(f"File successfully created at: {output_path}")
print(f"Final Size: {file_size:.2f} KB")
