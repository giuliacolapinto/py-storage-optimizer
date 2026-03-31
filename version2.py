############################    STORAGE Algorithm     ###########################
# Version 2

import pandas as pd
from openpyxl import Workbook
import os

# --- CONFIGURATION ---
input_path = '/set/your/path/walmartSales.xlsx'
output_path = "/set/your/path/walmart_version2.xlsx"

# Base 52 Mapping (a-z, A-Z)
MAP = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"

def to_base_52(n):
    if n == 1: return "" 
    if n <= 52: return MAP[n-1]
    # For counts > 52, it uses two letters
    return MAP[(n // 52) - 1] + MAP[(n % 52) - 1]

# --- PROCESSING ---
try:
    df = pd.read_excel(input_path, sheet_name='Walmart sales')
except Exception as e:
    print(f"Error loading file: {e}")
    exit()

wb = Workbook()
ws = wb.active
ws.title = "W"

for index, row in enumerate(df.values):
    id_v = row[0]
    # Ensure row data is treated as string before splitting
    data_list = [int(v) for v in str(row[1]).split()]
    
    compressed = ""
    i = 0
    while i < len(data_list):
        val = data_list[i]
        count = 1
        # Count consecutive identical values
        while i + count < len(data_list) and data_list[i + count] == val:
            count += 1
         
        # Format: Value + Base52_Letter_Code
        letter_code = to_base_52(count)
        compressed += f"{val}{letter_code}"
        i += count

    ws.cell(row=index + 1, column=1, value=id_v)
    ws.cell(row=index + 1, column=2, value=compressed)

# --- SAVE ---
# If file exists, it will be overwritten
wb.save(output_path)
print(f"File successfully created at: {output_path}")